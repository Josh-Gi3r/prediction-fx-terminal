#!/usr/bin/env python3
"""
Parse WC2026_Prediction_Market_Master.xlsx into lib/wc2026/data.ts.

Usage:
    python3 scripts/parse-wc2026.py

Reads:
    _reference/WC2026_Prediction_Market_Master.xlsx
Writes:
    lib/wc2026/data.ts
"""
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "_reference" / "WC2026_Prediction_Market_Master.xlsx"
OUT = ROOT / "lib" / "wc2026" / "data.ts"


def cell(v):
    """Normalize cell values to JSON-safe scalars (None becomes null)."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s == "—" or s == "-":
            return None
        return s
    return v


def parse_matches(ws):
    out = []
    # Skip title (row 1) + header (row 2)
    for row in ws.iter_rows(min_row=3, values_only=True):
        m = row[1]
        if m is None:
            continue
        out.append({
            "matchNumber": int(m),
            "stage": cell(row[2]) or "",
            "date": cell(row[3]) or "",
            "group": cell(row[4]),
            "homeTeam": cell(row[5]) or "",
            "awayTeam": cell(row[7]) or "",
            "venue": cell(row[8]) or "",
            "moneylineHome": cell(row[9]),
            "drawTie": cell(row[10]),
            "moneylineAway": cell(row[11]),
            "overUnderGoals": cell(row[12]),
        })
    return out


def parse_outright(ws):
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        team = cell(row[1])
        if not team:
            continue
        out.append({
            "team": team,
            "group": cell(row[2]),
            "draftkings": cell(row[3]),
            "betmgm": cell(row[4]),
            "fanduel": cell(row[5]),
            "bet365": cell(row[6]),
            "sportsbetAU": cell(row[7]),
            "williamHill": cell(row[8]),
            "polymarketPct": cell(row[9]),
            "kalshiImplied": cell(row[10]),
        })
    return out


def parse_knockout(ws):
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        team = cell(row[1])
        if not team:
            continue
        out.append({
            "team": team,
            "group": cell(row[2]),
            "toWin": cell(row[3]),
            "toReachFinal": cell(row[4]),
            "toReachSemis": cell(row[5]),
            "toReachQF": cell(row[6]),
            "toReachR16": cell(row[7]),
            "outInGroups": cell(row[8]),
            "draftkingsWin": cell(row[9]),
            "polymarketWinPct": cell(row[10]),
        })
    return out


def parse_group_odds(ws):
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        group = cell(row[1])
        team = cell(row[2])
        if not group or not team:
            continue
        out.append({
            "group": group,
            "team": team,
            "fanduelWinGroup": cell(row[3]),
            "polymarketWinPct": cell(row[4]),
            "toQualifyTop2": cell(row[5]),
            "draftkingsQualifyPct": cell(row[6]),
            "thirdPlaceAdvancePct": cell(row[7]),
            "eliminatedInGroupsPct": cell(row[8]),
        })
    return out


def parse_player_odds(ws):
    """Golden Boot sheet: subsections separated by section header rows."""
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        player = cell(row[1])
        nation = cell(row[2])
        # Player rows are the ones with both a name and a nation
        if not player or not nation or player in {"Player", "Goalkeeper"}:
            continue
        out.append({
            "player": player,
            "nation": nation,
            "position": cell(row[3]) or "",
            "club": cell(row[4]) or "",
            "draftkings": cell(row[5]),
            "sportsbetAU": cell(row[6]),
            "oddscheckerUK": cell(row[7]),
            "kalshiImplied": cell(row[8]),
        })
    return out


def parse_prediction_markets(ws):
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        cat = cell(row[1])
        name = cell(row[2])
        if not cat or not name:
            continue
        out.append({
            "category": cat,
            "marketName": name,
            "platform": cell(row[3]) or "",
            "currentOdds": cell(row[4]) or "",
            "volume": cell(row[5]),
            "notes": cell(row[6]),
        })
    return out


VALID_GROUPS = set("ABCDEFGHIJKL")


def parse_squads(ws):
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        group = cell(row[1])
        team = cell(row[2])
        name = cell(row[5])
        if not group or not team or not name:
            continue
        # Skip footer/note rows like "ALL — Remaining 40 teams"
        if group not in VALID_GROUPS:
            continue
        if isinstance(team, str) and team.upper().startswith("NOTE"):
            continue
        jersey = row[3]
        age = row[6]
        caps = row[7]
        goals = row[8]
        out.append({
            "group": group,
            "team": team,
            "jersey": int(jersey) if isinstance(jersey, (int, float)) else None,
            "position": cell(row[4]) or "",
            "playerName": name,
            "age": int(age) if isinstance(age, (int, float)) else None,
            "caps": int(caps) if isinstance(caps, (int, float)) else None,
            "goals": int(goals) if isinstance(goals, (int, float)) else None,
            "club": cell(row[9]) or "",
        })
    return out


def emit_ts(matches, outright, knockout, group_odds, players, markets, squads):
    """Emit a TS module with as-const literals."""
    def fmt(rows, type_name):
        body = json.dumps(rows, ensure_ascii=False, indent=2)
        return f"export const __DATA__: readonly {type_name}[] = {body} as const;\n"

    parts = [
        "// AUTO-GENERATED by scripts/parse-wc2026.py — do not edit by hand.",
        "// Source: _reference/WC2026_Prediction_Market_Master.xlsx",
        "// Regenerate: bun run data:wc2026",
        "",
        'import type {',
        '  GroupOdds,',
        '  KnockoutOdds,',
        '  Match,',
        '  OutrightOdds,',
        '  PlayerOdds,',
        '  PredictionMarket,',
        '  SquadPlayer,',
        '} from "./types";',
        "",
        f"export const MATCHES: readonly Match[] = {json.dumps(matches, ensure_ascii=False, indent=2)};",
        "",
        f"export const OUTRIGHT_ODDS: readonly OutrightOdds[] = {json.dumps(outright, ensure_ascii=False, indent=2)};",
        "",
        f"export const KNOCKOUT_ODDS: readonly KnockoutOdds[] = {json.dumps(knockout, ensure_ascii=False, indent=2)};",
        "",
        f"export const GROUP_ODDS: readonly GroupOdds[] = {json.dumps(group_odds, ensure_ascii=False, indent=2)};",
        "",
        f"export const PLAYER_ODDS: readonly PlayerOdds[] = {json.dumps(players, ensure_ascii=False, indent=2)};",
        "",
        f"export const PREDICTION_MARKETS: readonly PredictionMarket[] = {json.dumps(markets, ensure_ascii=False, indent=2)};",
        "",
        f"export const SQUADS: readonly SquadPlayer[] = {json.dumps(squads, ensure_ascii=False, indent=2)};",
        "",
        "/** All 12 groups derived from OUTRIGHT_ODDS. */",
        'export const GROUPS = ["A","B","C","D","E","F","G","H","I","J","K","L"] as const;',
        "",
    ]
    return "\n".join(parts)


def main():
    print(f"Reading {XLSX.relative_to(ROOT)}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    matches = parse_matches(wb["Match Schedule"])
    outright = parse_outright(wb["Outright Winner Odds"])
    knockout = parse_knockout(wb["Knockout Odds"])
    group_odds = parse_group_odds(wb["Group Odds"])
    players = parse_player_odds(wb["Golden Boot & Awards"])
    markets = parse_prediction_markets(wb["Prediction Markets"])
    squads = parse_squads(wb["Squads — All 48 Teams"])

    print(f"  Matches:         {len(matches)}")
    print(f"  Outright odds:   {len(outright)}")
    print(f"  Knockout odds:   {len(knockout)}")
    print(f"  Group odds:      {len(group_odds)}")
    print(f"  Player odds:     {len(players)}")
    print(f"  Markets:         {len(markets)}")
    print(f"  Squad players:   {len(squads)}")

    ts = emit_ts(matches, outright, knockout, group_odds, players, markets, squads)
    OUT.write_text(ts, encoding="utf-8")
    print(f"Wrote {OUT.relative_to(ROOT)} ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
